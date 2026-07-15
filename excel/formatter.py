"""Shared worksheet formatting so all output sheets look consistent."""

from __future__ import annotations

from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

_MAX_COLUMN_WIDTH = 60
_COLUMN_PADDING = 2


def apply_standard_formatting(worksheet: Worksheet) -> None:
    """Apply bold headers, a frozen header row, and auto-sized columns."""
    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    worksheet.freeze_panes = "A2"
    _auto_size_columns(worksheet)


def _auto_size_columns(worksheet: Worksheet) -> None:
    column_widths: dict[str, int] = {}
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            column_widths[cell.column_letter] = max(column_widths.get(cell.column_letter, 0), length)

    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = min(width + _COLUMN_PADDING, _MAX_COLUMN_WIDTH)
